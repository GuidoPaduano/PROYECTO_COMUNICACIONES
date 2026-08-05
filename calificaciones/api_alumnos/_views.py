# calificaciones/api_alumnos/_views.py
import io

from django.conf import settings
from django.http import HttpResponse
from django.db import IntegrityError, transaction
from django.views.decorators.csrf import csrf_exempt

from rest_framework.decorators import (
    api_view,
    authentication_classes,
    permission_classes,
    parser_classes,
)
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.parsers import JSONParser, FormParser, MultiPartParser

from ..jwt_auth import CookieJWTAuthentication as JWTAuthentication
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group

from ..models import Alumno, SchoolCourse
from ..models_preceptores import SchoolMembership
from ..schools import get_request_school, get_school_by_identifier, school_to_dict
from ..user_groups import get_user_group_names
from ..utils_cursos import (
    VALID_CURSOS,
    clear_school_course_cache,
    get_school_course_dicts,
    resolve_course_reference,
)

from ._helpers import (
    _alumno_base_qs,
    _alumno_matches_target_course,
    _alumno_to_dict,
    _build_import_plan,
    _can_manage_alumnos,
    _course_code_for_storage,
    _generar_id_alumno_para_curso,
    _generar_password_aleatoria,
    _generar_username,
    _is_valid_curso,
    _legajo_exists_in_school,
    _parse_import_file,
    _preceptor_assignment_guard,
    _resolve_alumno_for_transfer,
    _truthy,
)
from ..views._acceso import (
    _has_role,
    _preceptor_assignment_refs,
    _profesor_assignment_refs,
    _school_course_options_for_ui,
)
from ..course_access import filter_course_options_by_refs


@csrf_exempt
@api_view(["GET"])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def admin_importar_alumnos_template(request):
    if not getattr(request.user, "is_superuser", False):
        return Response({"detail": "No autorizado."}, status=403)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception:
        return Response({"detail": "No se pudo generar la plantilla Excel."}, status=500)

    school_ref = (
        request.query_params.get("school")
        or request.headers.get("X-School")
        or ""
    )
    school = get_school_by_identifier(school_ref) if school_ref else None
    if school is not None:
        cursos = list(
            SchoolCourse.objects.filter(school=school, is_active=True)
            .order_by("sort_order", "code", "id")
            .values_list("code", flat=True)
        )
    else:
        cursos = list(VALID_CURSOS)

    workbook = Workbook()
    header_fill = PatternFill(fill_type="solid", fgColor="E8EEF9")
    course_fill = PatternFill(fill_type="solid", fgColor="F0F4FF")
    headers = [
        "Curso",
        "Apellido Estudiante",
        "Nombre Estudiante",
        "Apellido Padre/Madre/Tutor",
        "Nombre Padre/Madre/Tutor",
        "Mail",
    ]
    col_widths = [12, 28, 24, 32, 28, 36]

    sheet = workbook.active
    sheet.title = "Alumnos"
    sheet.append(headers)
    sheet.freeze_panes = "A2"
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        sheet.column_dimensions[get_column_letter(col_idx)].width = width

    for code in cursos:
        row_data = [str(code)] + [""] * (len(headers) - 1)
        sheet.append(row_data)
        curso_cell = sheet.cell(row=sheet.max_row, column=1)
        curso_cell.fill = course_fill

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="plantilla-importacion-alumnos.xlsx"'
    return response


@csrf_exempt
@api_view(["POST"])
@parser_classes([JSONParser])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def admin_exportar_credenciales_xlsx(request):
    if not getattr(request.user, "is_superuser", False):
        return Response({"detail": "No autorizado."}, status=403)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception:
        return Response({"detail": "No se pudo generar el archivo Excel."}, status=500)

    credentials = request.data if isinstance(request.data, list) else []

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Credenciales"

    header_fill = PatternFill(fill_type="solid", fgColor="E8EEF9")
    headers = [
        "Legajo", "Apellido", "Nombre", "Curso",
        "Usuario Alumno", "Contraseña Alumno",
        "Nombre Padre/Tutor", "Apellido Padre/Tutor",
        "Mail Padre/Tutor", "Contraseña Padre/Tutor",
    ]
    col_widths = [14, 22, 22, 10, 26, 18, 22, 22, 36, 18]

    sheet.append(headers)
    sheet.freeze_panes = "A2"
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        sheet.column_dimensions[get_column_letter(col_idx)].width = width

    for item in credentials:
        sheet.append([
            item.get("legajo", ""),
            item.get("apellido", ""),
            item.get("nombre", ""),
            item.get("curso", ""),
            item.get("usuario_alumno", ""),
            str(item.get("password_alumno", "") or ""),
            item.get("nombre_padre", ""),
            item.get("apellido_padre", ""),
            item.get("mail_padre", ""),
            str(item.get("password_padre", "") or ""),
        ])

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="credenciales-importacion.xlsx"'
    return response


@csrf_exempt
@api_view(["POST"])
@parser_classes([JSONParser])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def crear_alumno(request):
    """
    POST /alumnos/crear/
    JSON:
      {
        "id_alumno": "A00123",    # opcional (si no viene, se genera)
        "nombre": "Luca",         # requerido (si no viene, se usa el id_alumno)
        "apellido": "Cabrera",    # opcional ("" por defecto)
        "school_course_id": 14    # requerido
      }
    """
    data = request.data or {}
    # Soportamos "legajo" como alias
    id_alumno = (data.get("id_alumno") or data.get("legajo") or "").strip()
    nombre = (data.get("nombre") or "").strip()
    apellido = (data.get("apellido") or "").strip()
    active_school = get_request_school(request)
    school_course, curso, course_error = resolve_course_reference(
        school=active_school,
        raw_course=data.get("curso"),
        raw_school_course_id=data.get("school_course_id"),
        required=True,
    )
    user_supplied_legajo = bool(id_alumno)

    if not _can_manage_alumnos(request.user, school=active_school):
        return Response({"detail": "No autorizado."}, status=403)

    if course_error:
        return Response({"detail": course_error}, status=400)
    course_code = _course_code_for_storage(school_course=school_course, curso=curso)
    if not _is_valid_curso(course_code, school=active_school):
        return Response({"detail": f"Curso inválido: {course_code}."}, status=400)
    if active_school is not None and school_course is None:
        return Response({"detail": "No existe ese curso en el colegio activo."}, status=400)
    guard_response = _preceptor_assignment_guard(
        user=request.user,
        school=active_school,
        school_course=school_course,
        curso=course_code,
        target_detail="No autorizado para crear alumnos en ese curso.",
    )
    if guard_response is not None:
        return guard_response

    # id_alumno ahora es realmente opcional (si no viene, lo generamos)
    if not id_alumno:
        id_alumno = _generar_id_alumno_para_curso(course_code, school=active_school, school_course=school_course)
    elif _legajo_exists_in_school(id_alumno, school=active_school):
        return Response({"detail": "El id_alumno (legajo) ya existe en este colegio."}, status=400)

    # El modelo requiere nombre; si el frontend manda solo legajo, ponemos algo razonable.
    if not nombre:
        nombre = id_alumno

    # Crear de forma segura (por si justo colisiona el legajo generado)
    try:
        with transaction.atomic():
            a = Alumno.objects.create(
                id_alumno=id_alumno,
                nombre=nombre,
                apellido=apellido,
                school=active_school,
                school_course=school_course,
                curso=course_code,
                padre=None,  # opcional; se puede asociar luego
            )
    except IntegrityError:
        # Si el usuario lo escribió y chocó, devolvemos el error claro.
        # Si lo generamos y chocó por carrera, reintentamos 1 vez.
        if user_supplied_legajo:
            return Response({"detail": "El id_alumno (legajo) ya existe en este colegio."}, status=400)

        try:
            id_alumno2 = _generar_id_alumno_para_curso(
                course_code,
                school=active_school,
                school_course=school_course,
            )
            with transaction.atomic():
                a = Alumno.objects.create(
                    id_alumno=id_alumno2,
                    nombre=nombre,
                    apellido=apellido,
                    school=active_school,
                    school_course=school_course,
                    curso=course_code,
                    padre=None,
                )
        except IntegrityError:
            return Response({"detail": "No se pudo generar un id_alumno único."}, status=400)

    return Response(
        {
            "alumno": _alumno_to_dict(a),
        },
        status=201,
    )


@csrf_exempt
@api_view(["POST"])
@parser_classes([MultiPartParser, FormParser])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def admin_importar_alumnos(request):
    if not getattr(request.user, "is_superuser", False):
        return Response({"detail": "No autorizado."}, status=403)

    school_ref = (
        request.data.get("school")
        or request.data.get("school_id")
        or request.data.get("school_slug")
        or request.headers.get("X-School")
        or ""
    )
    school = get_school_by_identifier(school_ref)
    if school is None:
        return Response({"detail": "Seleccioná un colegio válido."}, status=400)

    uploaded = request.FILES.get("file")
    if uploaded is None:
        return Response({"detail": "Subí un archivo .xlsx o .csv."}, status=400)

    max_bytes = int(getattr(settings, "STUDENT_IMPORT_MAX_BYTES", 5 * 1024 * 1024))
    uploaded_size = int(getattr(uploaded, "size", 0) or 0)
    if uploaded_size > max_bytes:
        max_mb = max_bytes / (1024 * 1024)
        return Response(
            {"detail": f"El archivo supera el limite permitido de {max_mb:g} MB."},
            status=413,
        )

    commit = _truthy(request.data.get("commit"))
    try:
        rows = _parse_import_file(uploaded)
    except ValueError as exc:
        return Response({"detail": str(exc)}, status=400)

    plan, errors, skipped, courses_to_create = _build_import_plan(rows=rows, school=school)
    created = []
    created_courses = []

    if commit and errors:
        return Response(
            {
                "detail": "Corregí los errores antes de importar.",
                "school": school_to_dict(school),
                "summary": {
                    "valid": len(plan),
                    "errors": len(errors),
                    "skipped": len(skipped),
                    "created": 0,
                },
                "errors": errors[:100],
                "skipped": skipped[:100],
                "preview": [
                    {k: v for k, v in item.items() if k != "school_course"}
                    for item in plan[:100]
                ],
                "courses_to_create": courses_to_create[:100],
                "created_courses": [],
            },
            status=400,
        )

    credentials = []

    if commit:
        User = get_user_model()
        group_alumnos, _ = Group.objects.get_or_create(name="Alumnos")
        group_padres, _ = Group.objects.get_or_create(name="Padres")

        try:
            with transaction.atomic():
                padres_creados = {}  # mail → User, para reutilizar en mellizos/hermanos

                course_map = {
                    str(course.code or "").strip().upper(): course
                    for course in SchoolCourse.objects.filter(school=school, is_active=True)
                }
                next_sort_order = (
                    SchoolCourse.objects.filter(school=school)
                    .order_by("-sort_order")
                    .values_list("sort_order", flat=True)
                    .first()
                    or 0
                )
                for course_info in courses_to_create:
                    code = course_info["code"]
                    school_course = course_map.get(code)
                    if school_course is None:
                        next_sort_order += 1
                        school_course, was_created = SchoolCourse.objects.get_or_create(
                            school=school,
                            code=code,
                            defaults={
                                "name": course_info["name"],
                                "is_active": True,
                                "sort_order": next_sort_order,
                            },
                        )
                        if not was_created and not school_course.is_active:
                            school_course.is_active = True
                            school_course.save(update_fields=["is_active"])
                        if was_created:
                            created_courses.append(
                                {
                                    "id": school_course.id,
                                    "code": school_course.code,
                                    "name": school_course.name,
                                }
                            )
                    course_map[code] = school_course

                for item in plan:
                    school_course = item["school_course"] or course_map.get(item["curso"])
                    if school_course is None:
                        raise IntegrityError("No se pudo resolver el curso de la fila importada.")

                    # Usuario del alumno
                    username_alumno = _generar_username(item["nombre"], item["apellido"])
                    password_alumno = _generar_password_aleatoria()
                    user_alumno = User.objects.create_user(
                        username=username_alumno,
                        password=password_alumno,
                        first_name=item["nombre"],
                        last_name=item["apellido"],
                    )
                    user_alumno.groups.add(group_alumnos)
                    SchoolMembership.objects.get_or_create(school=school, user=user_alumno)

                    # Usuario del padre/tutor (si hay datos)
                    user_padre = None
                    password_padre = None
                    if item.get("has_padre_data") and item.get("mail_padre"):
                        mail_padre = item["mail_padre"]
                        if mail_padre in padres_creados:
                            # Mellizos/hermanos: reutilizar padre ya creado en esta importación
                            user_padre = padres_creados[mail_padre]
                        else:
                            # Buscar si el usuario ya existe en el sistema por username o email
                            user_padre = (
                                User.objects.filter(username=mail_padre).first()
                                or User.objects.filter(email=mail_padre).first()
                            )
                            if user_padre is None:
                                password_padre = _generar_password_aleatoria()
                                user_padre = User.objects.create_user(
                                    username=mail_padre,
                                    email=mail_padre,
                                    password=password_padre,
                                    first_name=item.get("nombre_padre", ""),
                                    last_name=item.get("apellido_padre", ""),
                                )
                                user_padre.groups.add(group_padres)
                            else:
                                # Verificar que no tenga un rol incompatible con ser padre
                                roles_existentes = set(
                                    user_padre.groups.values_list("name", flat=True)
                                )
                                roles_incompatibles = roles_existentes & {
                                    "Profesores", "Directivos", "Preceptores"
                                }
                                if roles_incompatibles or user_padre.is_staff or user_padre.is_superuser:
                                    errors.append({
                                        "fila": item.get("_fila", "?"),
                                        "error": (
                                            f"El email {mail_padre} ya pertenece a un usuario con rol "
                                            f"incompatible ({', '.join(roles_incompatibles) or 'staff/admin'}). "
                                            "No se puede vincular como padre."
                                        ),
                                    })
                                    user_padre = None
                                else:
                                    # Asegurar que tenga el grupo Padres
                                    if "Padres" not in roles_existentes:
                                        user_padre.groups.add(group_padres)
                            if user_padre is not None:
                                SchoolMembership.objects.get_or_create(school=school, user=user_padre)
                            padres_creados[mail_padre] = user_padre

                    alumno = Alumno.objects.create(
                        school=school,
                        school_course=school_course,
                        curso=item["curso"],
                        id_alumno=item["legajo"],
                        nombre=item["nombre"],
                        apellido=item["apellido"],
                        usuario=user_alumno,
                        padre=user_padre,
                    )
                    created.append(_alumno_to_dict(alumno))
                    credentials.append({
                        "legajo": item["legajo"],
                        "nombre": item["nombre"],
                        "apellido": item["apellido"],
                        "curso": item["curso"],
                        "usuario_alumno": username_alumno,
                        "password_alumno": password_alumno,
                        "nombre_padre": item.get("nombre_padre", ""),
                        "apellido_padre": item.get("apellido_padre", ""),
                        "mail_padre": item.get("mail_padre", ""),
                        "password_padre": password_padre,
                    })

                if created_courses:
                    clear_school_course_cache(school)
        except IntegrityError:
            return Response({"detail": "La importación encontró legajos duplicados. Volvé a previsualizar."}, status=400)

    return Response(
        {
            "school": school_to_dict(school),
            "summary": {
                "valid": len(plan),
                "errors": len(errors),
                "skipped": len(skipped),
                "created": len(created),
                "courses_to_create": len(courses_to_create),
                "created_courses": len(created_courses),
            },
            "errors": errors[:100],
            "skipped": skipped[:100],
            "preview": [
                {k: v for k, v in item.items() if k not in {"school_course"}}
                for item in plan[:100]
            ],
            "courses_to_create": courses_to_create[:100],
            "created_courses": created_courses[:100],
            "created": created[:100],
            "credentials": credentials,
        },
        status=201 if commit else 200,
    )


@csrf_exempt
@api_view(["POST"])
@parser_classes([JSONParser])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def vincular_mi_legajo(request):
    """
    POST /alumnos/vincular/
    Vincula el usuario autenticado al registro Alumno (por legajo/id_alumno).
    Esto resuelve de forma explícita el problema de "no pudimos determinar tu alumno".

    JSON:
      {
        "id_alumno": "A00123"
      }

    Reglas:
    - Busca Alumno por id_alumno (case-insensitive).
    - Si el Alumno ya está vinculado a OTRO usuario -> 409.
    - Si ya está vinculado al mismo usuario -> 200 (idempotente).
    - Si está libre -> vincula y devuelve 200.
    """
    data = request.data or {}
    id_alumno = (data.get("id_alumno") or data.get("legajo") or "").strip()
    active_school = get_request_school(request)
    groups = set(get_user_group_names(request.user))

    if not id_alumno:
        return Response({"detail": "Falta id_alumno (legajo)."}, status=400)
    if not getattr(request.user, "is_superuser", False) and not {
        "Alumnos",
        "Alumno",
    }.intersection(groups):
        return Response({"detail": "Solo un usuario Alumno puede vincular un legajo."}, status=403)
    if active_school is None:
        return Response({"detail": "No se pudo determinar el colegio activo."}, status=403)

    qs = _alumno_base_qs(active_school).filter(id_alumno__iexact=id_alumno)
    a = qs.first()
    if not a:
        return Response({"detail": "No existe un alumno con ese id_alumno (legajo)."}, status=404)

    # Si tu modelo todavía no tiene campo `usuario`, esto fallaría al intentar vincular.
    # Preferimos devolver un error explícito y útil.
    if not hasattr(a, "usuario_id"):
        return Response(
            {
                "detail": "El modelo Alumno no tiene el campo 'usuario'. "
                          "Necesitás agregarlo (OneToOne/FK) para poder vincular alumno↔usuario."
            },
            status=500,
        )

    # Conflicto: ya vinculado a otro usuario
    if a.usuario_id and a.usuario_id != request.user.id:
        return Response(
            {
                "detail": "Este alumno ya está vinculado a otro usuario.",
                "alumno": _alumno_to_dict(a),
            },
            status=409,
        )

    # Idempotente: ya está vinculado a este usuario
    if a.usuario_id == request.user.id:
        return Response(
            {
                "already_linked": True,
                "alumno": _alumno_to_dict(a),
            },
            status=200,
        )

    # Vincular
    a.usuario = request.user
    a.save(update_fields=["usuario"])

    return Response(
        {
            "already_linked": False,
            "alumno": _alumno_to_dict(a),
        },
        status=200,
    )


@csrf_exempt
@api_view(["GET"])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def cursos_disponibles(request):
    """
    GET /alumnos/cursos/
    Devuelve los cursos disponibles. Para preceptores y profesores filtra
    por sus cursos asignados; superusers y directivos ven todos.
    """
    active_school = get_request_school(request)
    user = request.user

    options = _school_course_options_for_ui(school=active_school)

    if not user.is_superuser and not _has_role(request, "Directivos"):
        if _has_role(request, "Preceptores"):
            refs = _preceptor_assignment_refs(user, school=active_school)
            options = filter_course_options_by_refs(options, refs) if refs else []
        elif _has_role(request, "Profesores"):
            refs = _profesor_assignment_refs(user, school=active_school)
            options = filter_course_options_by_refs(options, refs) if refs else []

    cursos = [
        {
            "id": opt.get("code") or opt.get("id"),
            "nombre": opt.get("label") or opt.get("nombre"),
            "code": opt.get("code") or opt.get("id"),
            "school_course_id": opt.get("school_course_id"),
        }
        for opt in options
        if opt.get("code") or opt.get("id")
    ]
    return Response({"cursos": cursos}, status=200)


@csrf_exempt
@api_view(["POST"])
@parser_classes([JSONParser])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def transferir_alumno(request):
    """
    POST /alumnos/transferir/
    JSON:
      {
        "alumno_id": 123,   # opcional si envias id_alumno
        "id_alumno": "1A002",
        "school_course_id": 22
      }
    """
    data = request.data or {}
    active_school = get_request_school(request)
    school_course, curso, course_error = resolve_course_reference(
        school=active_school,
        raw_course=data.get("curso"),
        raw_school_course_id=data.get("school_course_id"),
        required=True,
    )

    if not _can_manage_alumnos(request.user, school=active_school):
        return Response({"detail": "No autorizado."}, status=403)

    if course_error:
        return Response({"detail": course_error}, status=400)
    course_code = _course_code_for_storage(school_course=school_course, curso=curso)
    if not _is_valid_curso(course_code, school=active_school):
        return Response({"detail": f"Curso inválido: {course_code}."}, status=400)
    if active_school is not None and school_course is None:
        return Response({"detail": "No existe ese curso en el colegio activo."}, status=400)

    alumno = _resolve_alumno_for_transfer(data, school=active_school)
    if not alumno:
        return Response({"detail": "Alumno no encontrado."}, status=404)

    guard_response = _preceptor_assignment_guard(
        user=request.user,
        school=active_school,
        school_course=school_course,
        curso=course_code,
        alumno=alumno,
        current_detail="No autorizado para ese alumno.",
        target_detail="No autorizado para transferir al curso destino.",
    )
    if guard_response is not None:
        return guard_response

    if _alumno_matches_target_course(alumno, school_course=school_course, curso=course_code):
        return Response(
            {
                "alumno": _alumno_to_dict(alumno),
                "message": "El alumno ya pertenece a ese curso.",
            },
            status=200,
        )

    alumno.curso = course_code
    alumno.school_course = school_course
    alumno.save(update_fields=["curso", "school_course"])
    try:
        from ..api_padres import invalidate_mis_hijos_cache

        invalidate_mis_hijos_cache(
            user_id=getattr(alumno, "padre_id", None),
            school_id=getattr(alumno, "school_id", None),
        )
    except Exception:
        pass

    return Response(
        {
            "alumno": _alumno_to_dict(alumno),
        },
        status=200,
    )
