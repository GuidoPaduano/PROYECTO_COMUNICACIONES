# calificaciones/api_alumnos/_helpers.py
import csv
import io
import re
import unicodedata

from ..course_access import build_course_ref, course_ref_matches, get_assignment_course_refs
from ..models import Alumno, School, SchoolCourse
from ..schools import scope_queryset_to_school
from ..user_groups import get_user_group_names
from ..utils_cursos import (
    is_curso_valido,
    resolve_course_reference,
)

try:
    from ..models_preceptores import PreceptorCurso, SchoolAdmin  # type: ignore
except Exception:
    PreceptorCurso = None
    SchoolAdmin = None


def _is_valid_curso(curso: str, school=None) -> bool:
    return is_curso_valido(curso, school=school)


def _alumno_base_qs(school=None):
    return scope_queryset_to_school(
        Alumno.objects.select_related("school", "school_course", "padre", "usuario"),
        school,
    )


def _alumno_to_dict(a: Alumno) -> dict:
    # Mantengo el formato que ya venías usando en crear_alumno
    return {
        "id": a.id,
        "id_alumno": a.id_alumno,
        "nombre": getattr(a, "nombre", None),
        "apellido": getattr(a, "apellido", None),
        "school_id": getattr(a, "school_id", None),
        "school_course_id": getattr(a, "school_course_id", None),
        "school_course_name": getattr(getattr(a, "school_course", None), "name", None)
        or getattr(getattr(a, "school_course", None), "code", None)
        or getattr(a, "curso", None),
        "padre": getattr(a, "padre_id", None),
        # Si existe el campo usuario (OneToOne/FK), lo exponemos como id (si no existe, queda None)
        "usuario": getattr(a, "usuario_id", None) if hasattr(a, "usuario_id") else None,
    }


def _normalizar_prefijo_curso(curso: str) -> str:
    # Solo letras/números para que el legajo quede "limpio" (ej: 1A, 4NAT)
    return re.sub(r"[^A-Za-z0-9]", "", str(curso or "")).upper()


def _course_code_for_storage(*, school_course=None, curso: str = "", alumno: Alumno | None = None) -> str:
    alumno_school_course = getattr(alumno, "school_course", None) if alumno is not None else None
    return str(
        getattr(school_course, "code", None)
        or getattr(alumno_school_course, "code", None)
        or getattr(alumno, "curso", None)
        or curso
        or ""
    ).strip()


def _generar_id_alumno_para_curso(curso: str = "", school=None, school_course=None) -> str:
    """
    Genera un id_alumno único si el frontend lo deja vacío (campo "opcional").

    Formato: <CURSO><NNN>
      - 1A001, 1A002...
      - 4NAT001...

    Nota: no depende de DB-specific functions; trae los existentes y resuelve en Python.
    """
    pref = _normalizar_prefijo_curso(_course_code_for_storage(school_course=school_course, curso=curso))
    if not pref:
        pref = "AL"

    existentes = list(
        scope_queryset_to_school(
            Alumno.objects.filter(id_alumno__istartswith=pref),
            school,
        ).values_list("id_alumno", flat=True)
    )

    upper = {str(x).upper() for x in existentes if x}

    max_n = 0
    rgx = re.compile(rf"^{re.escape(pref)}(\\d+)$", re.IGNORECASE)
    for s in upper:
        m = rgx.match(s)
        if m:
            try:
                max_n = max(max_n, int(m.group(1)))
            except Exception:
                pass

    n = max_n + 1
    # Seguridad: por si hay huecos o id raros; igual debería cortar rápido.
    for _ in range(1, 5000):
        cand = f"{pref}{n:03d}"
        if cand.upper() not in upper:
            return cand
        n += 1

    # Último recurso (muy improbable)
    return f"{pref}{max_n + 1:03d}"


def _is_preceptor_user(user) -> bool:
    try:
        if user is None:
            return False
        groups = set(get_user_group_names(user))
        return bool({"Preceptores", "Preceptor", "Directivos", "Directivo"}.intersection(groups))
    except Exception:
        return False


def _is_directivo_user(user) -> bool:
    try:
        if user is None:
            return False
        groups = set(get_user_group_names(user))
        return "Directivos" in groups or "Directivo" in groups
    except Exception:
        return False


def _is_preceptor_course_scoped_user(user) -> bool:
    try:
        if user is None:
            return False
        groups = set(get_user_group_names(user))
        return "Preceptores" in groups or "Preceptor" in groups
    except Exception:
        return False


def _is_school_admin_for(user, school=None) -> bool:
    if SchoolAdmin is None or user is None:
        return False
    try:
        qs = SchoolAdmin.objects.filter(admin=user)
        if school is not None:
            qs = qs.filter(school=school)
        return qs.exists()
    except Exception:
        return False


def _can_manage_alumnos(user, school=None) -> bool:
    return bool(
        getattr(user, "is_superuser", False)
        or _is_preceptor_user(user)
        or _is_school_admin_for(user, school=school)
    )


def _preceptor_assignment_refs(user, school=None):
    if PreceptorCurso is None:
        return []

    school_id = getattr(school, "id", None) or 0
    cache_attr = "_cached_alumnos_preceptor_refs_by_school"
    cached = getattr(user, cache_attr, None)
    if isinstance(cached, dict) and school_id in cached:
        return list(cached[school_id])

    try:
        qs = PreceptorCurso.objects.filter(preceptor=user)
        if school is not None:
            qs = scope_queryset_to_school(qs, school)
        refs = get_assignment_course_refs(qs)
    except Exception:
        refs = []

    try:
        if not isinstance(cached, dict):
            cached = {}
        cached[school_id] = tuple(refs)
        setattr(user, cache_attr, cached)
    except Exception:
        pass
    return refs


def _preceptor_assignment_guard(
    *,
    user,
    school=None,
    school_course=None,
    curso: str | None = None,
    alumno: Alumno | None = None,
    current_detail: str = "No autorizado para ese alumno.",
    target_detail: str = "No autorizado para ese curso.",
):
    from rest_framework.response import Response

    if PreceptorCurso is None:
        return None
    if getattr(user, "is_superuser", False) or _is_directivo_user(user):
        return None
    if not _is_preceptor_course_scoped_user(user):
        return None

    try:
        refs = _preceptor_assignment_refs(user, school=school)
        if alumno is not None and not course_ref_matches(refs, obj=alumno):
            return Response({"detail": current_detail}, status=403)
        if not course_ref_matches(
            refs,
            school=school,
            school_course_id=getattr(school_course, "id", None),
            course_code=curso,
        ):
            return Response({"detail": target_detail}, status=403)
    except Exception:
        detail = current_detail if alumno is not None else target_detail
        return Response({"detail": detail}, status=403)
    return None


def _legajo_exists_in_school(id_alumno: str, school=None) -> bool:
    legajo = str(id_alumno or "").strip()
    if not legajo:
        return False
    try:
        return scope_queryset_to_school(Alumno.objects.all(), school).filter(id_alumno__iexact=legajo).exists()
    except Exception:
        return False


def _resolve_alumno_for_transfer(data, school=None) -> Alumno | None:
    alumno_id = data.get("alumno_id") or data.get("id")
    legajo = data.get("id_alumno") or data.get("legajo")
    qs = _alumno_base_qs(school)

    if alumno_id:
        try:
            return qs.get(pk=int(alumno_id))
        except Exception:
            return None

    if legajo:
        try:
            return qs.get(id_alumno__iexact=str(legajo).strip())
        except Exception:
            return None

    return None


def _alumno_matches_target_course(alumno: Alumno, *, school_course=None, curso: str = "") -> bool:
    refs = [build_course_ref(obj=alumno)]
    return course_ref_matches(
        refs,
        school_course_id=getattr(school_course, "id", None),
        course_code=_course_code_for_storage(school_course=school_course, curso=curso),
    )


def _truthy(value) -> bool:
    return str(value or "").strip().lower() in {"1", "true", "yes", "si", "sí", "on"}


def _normalize_import_header(value) -> str:
    raw = str(value or "").strip().lower()
    replacements = {
        "á": "a",
        "é": "e",
        "í": "i",
        "ó": "o",
        "ú": "u",
        "ñ": "n",
    }
    for src, dst in replacements.items():
        raw = raw.replace(src, dst)
    return re.sub(r"[^a-z0-9]+", "_", raw).strip("_")


def _first_import_value(row: dict, *keys: str) -> str:
    for key in keys:
        value = row.get(key)
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def _normalize_import_duplicate_value(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def _import_duplicate_signature(*, nombre: str, apellido: str, legajo: str, curso: str) -> tuple[str, str, str, str]:
    return (
        _normalize_import_duplicate_value(nombre),
        _normalize_import_duplicate_value(apellido),
        _normalize_import_duplicate_value(legajo),
        _normalize_import_duplicate_value(curso),
    )


def _validate_import_person_name(value: str, field_label: str) -> list[str]:
    text = str(value or "").strip()
    if not text:
        return [f"Falta {field_label}."]
    if any(ch.isdigit() for ch in text):
        return [f"El {field_label} no puede contener números."]
    return []


IGNORED_IMPORT_SHEET_TITLES = {
    "todo",
    "todos",
    "all",
    "alumnos",
    "estudiantes",
    "resumen",
    "base",
    "instrucciones",
    "instruccion",
    "instructions",
    "hoja",
    "hoja1",
    "sheet",
    "sheet1",
}


def _course_code_for_import(value: str) -> str:
    raw = str(value or "").strip()
    folded = unicodedata.normalize("NFKD", raw).encode("ascii", "ignore").decode("ascii")
    code = re.sub(r"[^A-Za-z0-9]", "", folded).upper()
    return code[:20]


def _course_name_for_import(value: str) -> str:
    name = re.sub(r"\s+", " ", str(value or "").strip())
    return name[:120]


def _course_from_import_sheet_title(title: str) -> str:
    normalized = _normalize_import_header(title)
    if not normalized or normalized in IGNORED_IMPORT_SHEET_TITLES:
        return ""
    return _course_code_for_import(title)


def _parse_import_file(uploaded):
    filename = str(getattr(uploaded, "name", "") or "").lower()
    raw_rows = []

    if filename.endswith(".csv"):
        content = uploaded.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(content))
        for row in reader:
            raw_rows.append({_normalize_import_header(k): v for k, v in (row or {}).items()})
        return raw_rows

    if filename.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise ValueError("El servidor no tiene soporte para Excel .xlsx instalado.") from exc

        workbook = load_workbook(uploaded, read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            implicit_course = _course_from_import_sheet_title(sheet.title)
            implicit_course_name = _course_name_for_import(sheet.title) if implicit_course else ""
            rows = iter(sheet.iter_rows(values_only=True))
            headers = []
            for raw_headers in rows:
                candidate = [_normalize_import_header(value) for value in (raw_headers or [])]
                candidate_set = set(candidate)
                has_student_name = bool(
                    candidate_set.intersection({"nombre", "nombres", "name", "apellido", "apellidos", "last_name"})
                )
                has_course = bool(candidate_set.intersection({"curso", "course", "grado", "division", "school_course"}))
                has_identifier = bool(candidate_set.intersection({"id_alumno", "legajo", "matricula", "id", "dni"}))
                if has_student_name and (has_course or has_identifier or implicit_course):
                    headers = candidate
                    break
            if not any(headers):
                continue
            for values in rows:
                item = {}
                for index, header in enumerate(headers):
                    if not header:
                        continue
                    value = values[index] if index < len(values) else ""
                    item[header] = "" if value is None else str(value).strip()
                if implicit_course and not _first_import_value(item, "curso", "course", "grado", "division", "school_course"):
                    item["curso"] = implicit_course
                    item["curso_nombre"] = implicit_course_name
                raw_rows.append(item)
        return raw_rows

    raise ValueError("Formato no soportado. Subí un archivo .xlsx o .csv.")


def _build_import_plan(*, rows: list[dict], school: School):
    courses_by_code = {}
    for course in SchoolCourse.objects.filter(school=school, is_active=True).order_by("sort_order", "code", "id"):
        raw_code = str(course.code or "").strip().upper()
        if raw_code:
            courses_by_code[raw_code] = course
        normalized_code = _course_code_for_import(raw_code)
        if normalized_code:
            courses_by_code.setdefault(normalized_code, course)
    existing_legajos = {
        str(value or "").strip().upper()
        for value in Alumno.objects.filter(school=school).values_list("id_alumno", flat=True)
        if str(value or "").strip()
    }

    plan = []
    errors = []
    skipped = []
    seen_legajos = set()
    seen_rows = {}
    courses_to_create = {}

    for index, row in enumerate(rows, start=2):
        nombre = _first_import_value(row, "nombre", "name", "nombres")
        apellido = _first_import_value(row, "apellido", "apellidos", "last_name")
        legajo = _first_import_value(row, "id_alumno", "legajo", "matricula", "id", "dni")
        raw_curso = _first_import_value(row, "curso", "course", "grado", "division", "school_course")
        curso = _course_code_for_import(raw_curso)
        curso_nombre = _course_name_for_import(
            _first_import_value(row, "curso_nombre", "nombre_curso", "course_name", "school_course_name")
            or raw_curso
        )

        if not any([nombre, apellido, legajo, curso]):
            continue

        row_errors = []
        row_signature = _import_duplicate_signature(
            nombre=nombre,
            apellido=apellido,
            legajo=legajo,
            curso=curso,
        )
        first_duplicate_row = seen_rows.get(row_signature)
        duplicate_row = first_duplicate_row is not None
        if first_duplicate_row is not None:
            row_errors.append(f"Fila duplicada dentro del archivo. Ya aparece en la fila {first_duplicate_row}.")
        else:
            seen_rows[row_signature] = index

        row_errors.extend(_validate_import_person_name(nombre, "nombre"))
        row_errors.extend(_validate_import_person_name(apellido, "apellido"))
        if not curso:
            row_errors.append("Falta curso.")

        school_course = courses_by_code.get(curso)
        if curso and school_course is None:
            courses_to_create.setdefault(curso, {"code": curso, "name": curso_nombre or curso})

        if duplicate_row:
            errors.append(
                {
                    "row": index,
                    "legajo": legajo,
                    "nombre": nombre,
                    "apellido": apellido,
                    "curso": curso,
                    "errors": row_errors,
                }
            )
            continue

        generated_legajo = False
        if not legajo and curso:
            legajo = _generar_id_alumno_para_curso(curso, school=school, school_course=school_course)
            generated_legajo = True
            base_match = re.match(r"^(.*?)(\d+)$", legajo)
            while str(legajo or "").strip().upper() in seen_legajos:
                if not base_match:
                    legajo = f"{legajo}1"
                    base_match = re.match(r"^(.*?)(\d+)$", legajo)
                    continue
                prefix, number = base_match.groups()
                legajo = f"{prefix}{int(number) + 1:0{len(number)}d}"
                base_match = re.match(r"^(.*?)(\d+)$", legajo)

        legajo_key = str(legajo or "").strip().upper()
        if not legajo_key:
            row_errors.append("Falta legajo.")
        elif legajo_key in existing_legajos:
            skipped.append(
                {
                    "row": index,
                    "legajo": legajo,
                    "nombre": nombre,
                    "apellido": apellido,
                    "curso": curso,
                    "reason": "Ya existe un alumno con ese legajo en este colegio.",
                }
            )
            continue
        elif legajo_key in seen_legajos:
            row_errors.append("Legajo duplicado dentro del archivo.")

        if row_errors:
            errors.append(
                {
                    "row": index,
                    "legajo": legajo,
                    "nombre": nombre,
                    "apellido": apellido,
                    "curso": curso,
                    "errors": row_errors,
                }
            )
            continue

        seen_legajos.add(legajo_key)
        plan.append(
            {
                "row": index,
                "legajo": legajo,
                "nombre": nombre,
                "apellido": apellido,
                "curso": curso,
                "school_course": school_course,
                "school_course_id": getattr(school_course, "id", None),
                "school_course_name": getattr(school_course, "name", None) or curso_nombre or curso,
                "will_create_school_course": school_course is None,
                "generated_legajo": generated_legajo,
            }
        )

    return plan, errors, skipped, list(courses_to_create.values())
