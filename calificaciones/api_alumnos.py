# calificaciones/api_alumnos.py
import csv
import io
import re

from django.db import IntegrityError, transaction
from django.views.decorators.csrf import csrf_exempt

from rest_framework.decorators import (
    api_view,
    authentication_classes,
    permission_classes,
    parser_classes,
)
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.parsers import JSONParser, FormParser, MultiPartParser
from .jwt_auth import CookieJWTAuthentication as JWTAuthentication

from .course_access import build_course_ref, course_ref_matches, get_assignment_course_refs
from .models import Alumno, School, SchoolCourse
from .schools import get_request_school, get_school_by_identifier, school_to_dict, scope_queryset_to_school
from .user_groups import get_user_group_names
from .utils_cursos import get_school_course_dicts, is_curso_valido, resolve_course_reference

try:
    from .models_preceptores import PreceptorCurso  # type: ignore
except Exception:
    PreceptorCurso = None


def _is_valid_curso(curso: str, school=None) -> bool:
    return is_curso_valido(curso, school=school)


def _alumno_base_qs(school=None):
    return scope_queryset_to_school(
        Alumno.objects.select_related("school", "school_course", "padre", "usuario"),
        school,
    )


def _alumno_to_dict(a: Alumno) -> dict:
    # Mantengo el formato que ya venías usando en crear_alumno
    return {
        "id": a.id,
        "id_alumno": a.id_alumno,
        "nombre": getattr(a, "nombre", None),
        "apellido": getattr(a, "apellido", None),
        "school_id": getattr(a, "school_id", None),
        "school_course_id": getattr(a, "school_course_id", None),
        "school_course_name": getattr(getattr(a, "school_course", None), "name", None)
        or getattr(getattr(a, "school_course", None), "code", None)
        or getattr(a, "curso", None),
        "padre": getattr(a, "padre_id", None),
        # Si existe el campo usuario (OneToOne/FK), lo exponemos como id (si no existe, queda None)
        "usuario": getattr(a, "usuario_id", None) if hasattr(a, "usuario_id") else None,
    }


def _normalizar_prefijo_curso(curso: str) -> str:
    # Solo letras/números para que el legajo quede “limpio” (ej: 1A, 4NAT)
    return re.sub(r"[^A-Za-z0-9]", "", str(curso or "")).upper()


def _course_code_for_storage(*, school_course=None, curso: str = "", alumno: Alumno | None = None) -> str:
    alumno_school_course = getattr(alumno, "school_course", None) if alumno is not None else None
    return str(
        getattr(school_course, "code", None)
        or getattr(alumno_school_course, "code", None)
        or getattr(alumno, "curso", None)
        or curso
        or ""
    ).strip()


def _generar_id_alumno_para_curso(curso: str = "", school=None, school_course=None) -> str:
    """
    Genera un id_alumno único si el frontend lo deja vacío (campo "opcional").

    Formato: <CURSO><NNN>
      - 1A001, 1A002...
      - 4NAT001...

    Nota: no depende de DB-specific functions; trae los existentes y resuelve en Python.
    """
    pref = _normalizar_prefijo_curso(_course_code_for_storage(school_course=school_course, curso=curso))
    if not pref:
        pref = "AL"

    existentes = list(
        scope_queryset_to_school(
            Alumno.objects.filter(id_alumno__istartswith=pref),
            school,
        ).values_list("id_alumno", flat=True)
    )

    upper = {str(x).upper() for x in existentes if x}

    max_n = 0
    rgx = re.compile(rf"^{re.escape(pref)}(\\d+)$", re.IGNORECASE)
    for s in upper:
        m = rgx.match(s)
        if m:
            try:
                max_n = max(max_n, int(m.group(1)))
            except Exception:
                pass

    n = max_n + 1
    # Seguridad: por si hay huecos o id raros; igual debería cortar rápido.
    for _ in range(1, 5000):
        cand = f"{pref}{n:03d}"
        if cand.upper() not in upper:
            return cand
        n += 1

    # Último recurso (muy improbable)
    return f"{pref}{max_n + 1:03d}"


def _is_preceptor_user(user) -> bool:
    try:
        if user is None:
            return False
        groups = set(get_user_group_names(user))
        return bool({"Preceptores", "Preceptor", "Directivos", "Directivo"}.intersection(groups))
    except Exception:
        return False


def _is_directivo_user(user) -> bool:
    try:
        if user is None:
            return False
        groups = set(get_user_group_names(user))
        return "Directivos" in groups or "Directivo" in groups
    except Exception:
        return False


def _is_preceptor_course_scoped_user(user) -> bool:
    try:
        if user is None:
            return False
        groups = set(get_user_group_names(user))
        return "Preceptores" in groups or "Preceptor" in groups
    except Exception:
        return False


def _can_manage_alumnos(user) -> bool:
    return bool(getattr(user, "is_superuser", False) or _is_preceptor_user(user))


def _preceptor_assignment_refs(user, school=None):
    if PreceptorCurso is None:
        return []

    school_id = getattr(school, "id", None) or 0
    cache_attr = "_cached_alumnos_preceptor_refs_by_school"
    cached = getattr(user, cache_attr, None)
    if isinstance(cached, dict) and school_id in cached:
        return list(cached[school_id])

    try:
        qs = PreceptorCurso.objects.filter(preceptor=user)
        if school is not None:
            qs = scope_queryset_to_school(qs, school)
        refs = get_assignment_course_refs(qs)
    except Exception:
        refs = []

    try:
        if not isinstance(cached, dict):
            cached = {}
        cached[school_id] = tuple(refs)
        setattr(user, cache_attr, cached)
    except Exception:
        pass
    return refs


def _preceptor_assignment_guard(
    *,
    user,
    school=None,
    school_course=None,
    curso: str | None = None,
    alumno: Alumno | None = None,
    current_detail: str = "No autorizado para ese alumno.",
    target_detail: str = "No autorizado para ese curso.",
):
    if PreceptorCurso is None:
        return None
    if getattr(user, "is_superuser", False) or _is_directivo_user(user):
        return None
    if not _is_preceptor_course_scoped_user(user):
        return None

    try:
        refs = _preceptor_assignment_refs(user, school=school)
        if alumno is not None and not course_ref_matches(refs, obj=alumno):
            return Response({"detail": current_detail}, status=403)
        if not course_ref_matches(
            refs,
            school=school,
            school_course_id=getattr(school_course, "id", None),
            course_code=curso,
        ):
            return Response({"detail": target_detail}, status=403)
    except Exception:
        detail = current_detail if alumno is not None else target_detail
        return Response({"detail": detail}, status=403)
    return None


def _legajo_exists_in_school(id_alumno: str, school=None) -> bool:
    legajo = str(id_alumno or "").strip()
    if not legajo:
        return False
    try:
        return scope_queryset_to_school(Alumno.objects.all(), school).filter(id_alumno__iexact=legajo).exists()
    except Exception:
        return False


def _resolve_alumno_for_transfer(data, school=None) -> Alumno | None:
    alumno_id = data.get("alumno_id") or data.get("id")
    legajo = data.get("id_alumno") or data.get("legajo")
    qs = _alumno_base_qs(school)

    if alumno_id:
        try:
            return qs.get(pk=int(alumno_id))
        except Exception:
            return None

    if legajo:
        try:
            return qs.get(id_alumno__iexact=str(legajo).strip())
        except Exception:
            return None

    return None


def _alumno_matches_target_course(alumno: Alumno, *, school_course=None, curso: str = "") -> bool:
    refs = [build_course_ref(obj=alumno)]
    return course_ref_matches(
        refs,
        school_course_id=getattr(school_course, "id", None),
        course_code=_course_code_for_storage(school_course=school_course, curso=curso),
    )


@csrf_exempt
@api_view(["POST"])
@parser_classes([JSONParser])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def crear_alumno(request):
    """
    POST /alumnos/crear/
    JSON:
      {
        "id_alumno": "A00123",    # opcional (si no viene, se genera)
        "nombre": "Luca",         # requerido (si no viene, se usa el id_alumno)
        "apellido": "Cabrera",    # opcional ("" por defecto)
        "school_course_id": 14    # requerido
      }
    """
    data = request.data or {}
    # Soportamos "legajo" como alias
    id_alumno = (data.get("id_alumno") or data.get("legajo") or "").strip()
    nombre = (data.get("nombre") or "").strip()
    apellido = (data.get("apellido") or "").strip()
    active_school = get_request_school(request)
    school_course, curso, course_error = resolve_course_reference(
        school=active_school,
        raw_course=data.get("curso"),
        raw_school_course_id=data.get("school_course_id"),
        required=True,
    )
    user_supplied_legajo = bool(id_alumno)

    if not _can_manage_alumnos(request.user):
        return Response({"detail": "No autorizado."}, status=403)

    if course_error:
        return Response({"detail": course_error}, status=400)
    course_code = _course_code_for_storage(school_course=school_course, curso=curso)
    if not _is_valid_curso(course_code, school=active_school):
        return Response({"detail": f"Curso inválido: {course_code}."}, status=400)
    if active_school is not None and school_course is None:
        return Response({"detail": "No existe ese curso en el colegio activo."}, status=400)
    guard_response = _preceptor_assignment_guard(
        user=request.user,
        school=active_school,
        school_course=school_course,
        curso=course_code,
        target_detail="No autorizado para crear alumnos en ese curso.",
    )
    if guard_response is not None:
        return guard_response

    # ✅ id_alumno ahora es realmente opcional (si no viene, lo generamos)
    if not id_alumno:
        id_alumno = _generar_id_alumno_para_curso(course_code, school=active_school, school_course=school_course)
    elif _legajo_exists_in_school(id_alumno, school=active_school):
        return Response({"detail": "El id_alumno (legajo) ya existe en este colegio."}, status=400)

    # El modelo requiere nombre; si el frontend manda solo legajo, ponemos algo razonable.
    if not nombre:
        nombre = id_alumno

    # Crear de forma segura (por si justo colisiona el legajo generado)
    try:
        with transaction.atomic():
            a = Alumno.objects.create(
                id_alumno=id_alumno,
                nombre=nombre,
                apellido=apellido,
                school=active_school,
                school_course=school_course,
                curso=course_code,
                padre=None,  # opcional; se puede asociar luego
            )
    except IntegrityError:
        # Si el usuario lo escribió y chocó, devolvemos el error claro.
        # Si lo generamos y chocó por carrera, reintentamos 1 vez.
        if user_supplied_legajo:
            return Response({"detail": "El id_alumno (legajo) ya existe en este colegio."}, status=400)

        try:
            id_alumno2 = _generar_id_alumno_para_curso(
                course_code,
                school=active_school,
                school_course=school_course,
            )
            with transaction.atomic():
                a = Alumno.objects.create(
                    id_alumno=id_alumno2,
                    nombre=nombre,
                    apellido=apellido,
                    school=active_school,
                    school_course=school_course,
                    curso=course_code,
                    padre=None,
                )
        except IntegrityError:
            return Response({"detail": "No se pudo generar un id_alumno único."}, status=400)

    return Response(
        {
            "alumno": _alumno_to_dict(a),
        },
        status=201,
    )


def _truthy(value) -> bool:
    return str(value or "").strip().lower() in {"1", "true", "yes", "si", "sí", "on"}


def _normalize_import_header(value) -> str:
    raw = str(value or "").strip().lower()
    replacements = {
        "á": "a",
        "é": "e",
        "í": "i",
        "ó": "o",
        "ú": "u",
        "ñ": "n",
    }
    for src, dst in replacements.items():
        raw = raw.replace(src, dst)
    return re.sub(r"[^a-z0-9]+", "_", raw).strip("_")


def _first_import_value(row: dict, *keys: str) -> str:
    for key in keys:
        value = row.get(key)
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def _course_from_import_sheet_title(title: str) -> str:
    normalized = _normalize_import_header(title)
    if not normalized or normalized in {"todo", "todos", "all", "alumnos", "resumen", "base"}:
        return ""
    if re.search(r"\d", normalized):
        return re.sub(r"[^A-Za-z0-9]", "", str(title or "")).upper()
    return ""


def _parse_import_file(uploaded):
    filename = str(getattr(uploaded, "name", "") or "").lower()
    raw_rows = []

    if filename.endswith(".csv"):
        content = uploaded.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(content))
        for row in reader:
            raw_rows.append({_normalize_import_header(k): v for k, v in (row or {}).items()})
        return raw_rows

    if filename.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise ValueError("El servidor no tiene soporte para Excel .xlsx instalado.") from exc

        workbook = load_workbook(uploaded, read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            implicit_course = _course_from_import_sheet_title(sheet.title)
            rows = iter(sheet.iter_rows(values_only=True))
            headers = []
            for raw_headers in rows:
                candidate = [_normalize_import_header(value) for value in (raw_headers or [])]
                candidate_set = set(candidate)
                has_student_name = bool(
                    candidate_set.intersection({"nombre", "nombres", "name", "apellido", "apellidos", "last_name"})
                )
                has_course = bool(candidate_set.intersection({"curso", "course", "grado", "division", "school_course"}))
                has_identifier = bool(candidate_set.intersection({"id_alumno", "legajo", "matricula", "id", "dni"}))
                if has_student_name and (has_course or has_identifier or implicit_course):
                    headers = candidate
                    break
            if not any(headers):
                continue
            for values in rows:
                item = {}
                for index, header in enumerate(headers):
                    if not header:
                        continue
                    value = values[index] if index < len(values) else ""
                    item[header] = "" if value is None else str(value).strip()
                if implicit_course and not _first_import_value(item, "curso", "course", "grado", "division", "school_course"):
                    item["curso"] = implicit_course
                raw_rows.append(item)
        return raw_rows

    raise ValueError("Formato no soportado. Subí un archivo .xlsx o .csv.")


def _build_import_plan(*, rows: list[dict], school: School):
    courses_by_code = {
        str(course.code or "").strip().upper(): course
        for course in SchoolCourse.objects.filter(school=school, is_active=True).order_by("sort_order", "code", "id")
    }
    existing_legajos = {
        str(value or "").strip().upper()
        for value in Alumno.objects.filter(school=school).values_list("id_alumno", flat=True)
        if str(value or "").strip()
    }

    plan = []
    errors = []
    skipped = []
    seen_legajos = set()

    for index, row in enumerate(rows, start=2):
        nombre = _first_import_value(row, "nombre", "name", "nombres")
        apellido = _first_import_value(row, "apellido", "apellidos", "last_name")
        legajo = _first_import_value(row, "id_alumno", "legajo", "matricula", "id", "dni")
        curso = _first_import_value(row, "curso", "course", "grado", "division", "school_course").upper()

        if not any([nombre, apellido, legajo, curso]):
            continue

        row_errors = []
        if not nombre and not apellido:
            row_errors.append("Falta nombre o apellido.")
        if not curso:
            row_errors.append("Falta curso.")

        school_course = courses_by_code.get(curso)
        if curso and school_course is None:
            row_errors.append(f"Curso inexistente para {school.name}: {curso}.")

        generated_legajo = False
        if not legajo and school_course is not None:
            legajo = _generar_id_alumno_para_curso(curso, school=school, school_course=school_course)
            generated_legajo = True
            base_match = re.match(r"^(.*?)(\d+)$", legajo)
            while str(legajo or "").strip().upper() in seen_legajos:
                if not base_match:
                    legajo = f"{legajo}1"
                    base_match = re.match(r"^(.*?)(\d+)$", legajo)
                    continue
                prefix, number = base_match.groups()
                legajo = f"{prefix}{int(number) + 1:0{len(number)}d}"
                base_match = re.match(r"^(.*?)(\d+)$", legajo)

        legajo_key = str(legajo or "").strip().upper()
        if not legajo_key:
            row_errors.append("Falta legajo.")
        elif legajo_key in existing_legajos:
            skipped.append(
                {
                    "row": index,
                    "legajo": legajo,
                    "nombre": nombre,
                    "apellido": apellido,
                    "curso": curso,
                    "reason": "Ya existe un alumno con ese legajo en este colegio.",
                }
            )
            continue
        elif legajo_key in seen_legajos:
            row_errors.append("Legajo duplicado dentro del archivo.")

        if row_errors:
            errors.append(
                {
                    "row": index,
                    "legajo": legajo,
                    "nombre": nombre,
                    "apellido": apellido,
                    "curso": curso,
                    "errors": row_errors,
                }
            )
            continue

        seen_legajos.add(legajo_key)
        plan.append(
            {
                "row": index,
                "legajo": legajo,
                "nombre": nombre or legajo,
                "apellido": apellido,
                "curso": curso,
                "school_course": school_course,
                "school_course_id": school_course.id,
                "school_course_name": school_course.name,
                "generated_legajo": generated_legajo,
            }
        )

    return plan, errors, skipped


@csrf_exempt
@api_view(["POST"])
@parser_classes([MultiPartParser, FormParser])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def admin_importar_alumnos(request):
    if not getattr(request.user, "is_superuser", False):
        return Response({"detail": "No autorizado."}, status=403)

    school_ref = (
        request.data.get("school")
        or request.data.get("school_id")
        or request.data.get("school_slug")
        or request.headers.get("X-School")
        or ""
    )
    school = get_school_by_identifier(school_ref)
    if school is None:
        return Response({"detail": "Seleccioná un colegio válido."}, status=400)

    uploaded = request.FILES.get("file")
    if uploaded is None:
        return Response({"detail": "Subí un archivo .xlsx o .csv."}, status=400)

    commit = _truthy(request.data.get("commit"))
    try:
        rows = _parse_import_file(uploaded)
    except ValueError as exc:
        return Response({"detail": str(exc)}, status=400)

    plan, errors, skipped = _build_import_plan(rows=rows, school=school)
    created = []

    if commit and errors:
        return Response(
            {
                "detail": "Corregí los errores antes de importar.",
                "school": school_to_dict(school),
                "summary": {
                    "valid": len(plan),
                    "errors": len(errors),
                    "skipped": len(skipped),
                    "created": 0,
                },
                "errors": errors[:100],
                "skipped": skipped[:100],
                "preview": [
                    {k: v for k, v in item.items() if k != "school_course"}
                    for item in plan[:100]
                ],
            },
            status=400,
        )

    if commit:
        try:
            with transaction.atomic():
                for item in plan:
                    alumno = Alumno.objects.create(
                        school=school,
                        school_course=item["school_course"],
                        curso=item["curso"],
                        id_alumno=item["legajo"],
                        nombre=item["nombre"],
                        apellido=item["apellido"],
                    )
                    created.append(_alumno_to_dict(alumno))
        except IntegrityError:
            return Response({"detail": "La importación encontró legajos duplicados. Volvé a previsualizar."}, status=400)

    return Response(
        {
            "school": school_to_dict(school),
            "summary": {
                "valid": len(plan),
                "errors": len(errors),
                "skipped": len(skipped),
                "created": len(created),
            },
            "errors": errors[:100],
            "skipped": skipped[:100],
            "preview": [
                {k: v for k, v in item.items() if k != "school_course"}
                for item in plan[:100]
            ],
            "created": created[:100],
        },
        status=201 if commit else 200,
    )


@csrf_exempt
@api_view(["POST"])
@parser_classes([JSONParser])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def vincular_mi_legajo(request):
    """
    POST /alumnos/vincular/
    Vincula el usuario autenticado al registro Alumno (por legajo/id_alumno).
    Esto resuelve de forma explícita el problema de "no pudimos determinar tu alumno".

    JSON:
      {
        "id_alumno": "A00123"
      }

    Reglas:
    - Busca Alumno por id_alumno (case-insensitive).
    - Si el Alumno ya está vinculado a OTRO usuario -> 409.
    - Si ya está vinculado al mismo usuario -> 200 (idempotente).
    - Si está libre -> vincula y devuelve 200.
    """
    data = request.data or {}
    id_alumno = (data.get("id_alumno") or data.get("legajo") or "").strip()
    active_school = get_request_school(request)

    if not id_alumno:
        return Response({"detail": "Falta id_alumno (legajo)."}, status=400)

    qs = _alumno_base_qs(active_school).filter(id_alumno__iexact=id_alumno)
    a = qs.first()
    if not a:
        return Response({"detail": "No existe un alumno con ese id_alumno (legajo)."}, status=404)

    # Si tu modelo todavía no tiene campo `usuario`, esto fallaría al intentar vincular.
    # Preferimos devolver un error explícito y útil.
    if not hasattr(a, "usuario_id"):
        return Response(
            {
                "detail": "El modelo Alumno no tiene el campo 'usuario'. "
                          "Necesitás agregarlo (OneToOne/FK) para poder vincular alumno↔usuario."
            },
            status=500,
        )

    # Conflicto: ya vinculado a otro usuario
    if a.usuario_id and a.usuario_id != request.user.id:
        return Response(
            {
                "detail": "Este alumno ya está vinculado a otro usuario.",
                "alumno": _alumno_to_dict(a),
            },
            status=409,
        )

    # Idempotente: ya está vinculado a este usuario
    if a.usuario_id == request.user.id:
        return Response(
            {
                "already_linked": True,
                "alumno": _alumno_to_dict(a),
            },
            status=200,
        )

    # Vincular
    a.usuario = request.user
    a.save(update_fields=["usuario"])

    return Response(
        {
            "already_linked": False,
            "alumno": _alumno_to_dict(a),
        },
        status=200,
    )


@csrf_exempt
@api_view(["GET"])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def cursos_disponibles(request):
    """
    GET /alumnos/cursos/
    Devuelve todos los cursos disponibles (catalogo Alumno.CURSOS).
    """
    active_school = get_request_school(request)
    # Preferimos los cursos reales en DB (para no listar cursos inexistentes).
    # Si no hay alumnos, caemos al catálogo definido en el modelo.
    cursos = get_school_course_dicts(
        school=active_school,
        fallback_to_defaults=False,
        catalog_only=True,
    )
    cursos = [
        {
            "id": item.get("id"),
            "nombre": item.get("nombre"),
        }
        for item in cursos
        if item.get("id")
    ]
    return Response({"cursos": cursos}, status=200)


@csrf_exempt
@api_view(["POST"])
@parser_classes([JSONParser])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def transferir_alumno(request):
    """
    POST /alumnos/transferir/
    JSON:
      {
        "alumno_id": 123,   # opcional si envias id_alumno
        "id_alumno": "1A002",
        "school_course_id": 22
      }
    """
    data = request.data or {}
    active_school = get_request_school(request)
    school_course, curso, course_error = resolve_course_reference(
        school=active_school,
        raw_course=data.get("curso"),
        raw_school_course_id=data.get("school_course_id"),
        required=True,
    )

    if not _can_manage_alumnos(request.user):
        return Response({"detail": "No autorizado."}, status=403)

    if course_error:
        return Response({"detail": course_error}, status=400)
    course_code = _course_code_for_storage(school_course=school_course, curso=curso)
    if not _is_valid_curso(course_code, school=active_school):
        return Response({"detail": f"Curso inválido: {course_code}."}, status=400)
    if active_school is not None and school_course is None:
        return Response({"detail": "No existe ese curso en el colegio activo."}, status=400)

    alumno = _resolve_alumno_for_transfer(data, school=active_school)
    if not alumno:
        return Response({"detail": "Alumno no encontrado."}, status=404)

    guard_response = _preceptor_assignment_guard(
        user=request.user,
        school=active_school,
        school_course=school_course,
        curso=course_code,
        alumno=alumno,
        current_detail="No autorizado para ese alumno.",
        target_detail="No autorizado para transferir al curso destino.",
    )
    if guard_response is not None:
        return guard_response

    if _alumno_matches_target_course(alumno, school_course=school_course, curso=course_code):
        return Response(
            {
                "alumno": _alumno_to_dict(alumno),
                "message": "El alumno ya pertenece a ese curso.",
            },
            status=200,
        )

    alumno.curso = course_code
    alumno.school_course = school_course
    alumno.save(update_fields=["curso", "school_course"])

    return Response(
        {
            "alumno": _alumno_to_dict(alumno),
        },
        status=200,
    )
